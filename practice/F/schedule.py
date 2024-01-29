import os
import time
import datetime
import requests
import openpyxl
import pickle
import re
from enum import Enum
from collections import OrderedDict
from threading import Thread
from bs4 import BeautifulSoup


class Parity(Enum):
    even = 0
    odd = 1


class Weekdays(Enum):
    monday = 1
    tuesday = 2
    wednesday = 3
    thursday = 4
    friday = 5
    saturday = 6
    sunday = 7

    def from_string(weekday_string):
        for weekday in Weekdays:
            if str(weekday).upper() == weekday_string.upper():
                return weekday

    def __str__(self):
        weekdays_ru = [
            "Понедельник",
            "Вторник",
            "Среда",
            "Четверг",
            "Пятница",
            "Суббота",
            "Воскресенье",
        ]
        return weekdays_ru[self.value - 1]


class ScheduleTable:
    def __init__(self, pairs_per_day=7):
        self.pairs_per_day = pairs_per_day
        self.table = OrderedDict()
        for day in Weekdays:
            self.table[day] = OrderedDict()
            for pair in range(1, self.pairs_per_day + 1):
                self.table[day][pair] = OrderedDict()
                self.table[day][pair][Parity.odd] = None
                self.table[day][pair][Parity.even] = None

    def append(self, subject):
        for day in self.table:
            for pair in self.table[day]:
                for Parity in self.table[day][pair]:
                    if self.table[day][pair][Parity] is None:
                        self.table[day][pair][Parity] = subject
                        return (day, pair, Parity)

    def add(self, weekday, pair, Parity, subject):
        self.table[weekday][pair][Parity] = subject

    def get(self, weekday, pair, Parity):
        return self.table[weekday][pair][Parity]

    def __str__(self):
        return str(self.table)


class MireaSchedule(Thread):
    def __init__(self, update_time, data_dir="schedule_data"):
        Thread.__init__(self)
        self.daemon = True
        self.stopped = False

        self.groups_schedules_paths = dict()
        self.teachers_schedules_paths = dict()

        self.excel_tables_paths = list()
        self.data_dir = data_dir
        self.excel_tables_dir = os.path.join(self.data_dir, "excel_tables")
        self.pickle_tables_dir = os.path.join(self.data_dir, "pickle_tables")
        if not os.path.exists(self.data_dir):
            os.mkdir(self.data_dir)
        if not os.path.exists(self.excel_tables_dir):
            os.mkdir(self.excel_tables_dir)
        if not os.path.exists(self.pickle_tables_dir):
            os.mkdir(self.pickle_tables_dir)

        if type(update_time) == type(datetime.timedelta()):
            self.update_time = update_time.total_seconds()
        else:
            self.update_time = update_time

        print("Schedule: started")

    def get_schedule_tables_links(self):
        page = requests.get("https://mirea.ru/schedule")
        soup = BeautifulSoup(page.text, "html.parser")
        links = (
            soup.find("div", {"class": "schedule"})
            .find(string="Институт информационных технологий")
            .find_parent("div")
            .find_parent("div")
            .findAll("a", {"class": "uk-link-toggle"})
        )

        return [link["href"] for link in links]

    def download_schedule_tables(self):
        self.excel_tables_paths.clear()
        links = self.get_schedule_tables_links()
        for link in links:
            path = os.path.join(self.excel_tables_dir, link.split("/")[-1])
            with open(path, "wb") as table_file:
                res = requests.get(link)
                table_file.write(res.content)
                self.excel_tables_paths.append(path)

    def update_schedule(self):
        self.download_schedule_tables()
        groups_schedule = dict()
        teachers_schedule = dict()

        for excel_table_path in self.excel_tables_paths:
            if excel_table_path.split(os.sep)[-1][:3] == "IIT":
                table = openpyxl.load_workbook(excel_table_path)
                sheet = table.active

                group_name_pattern = r"^[А-Я]{4}-\d{2}-\d{2}$"
                teacher_name_pattern = r"[А-ЯЁ][а-яё]+\s[А-Я]\.[А-Я]\."
                base_row = 2
                last_row = 87
                for column in range(sheet.min_column, sheet.max_column + 1):
                    cell = str(sheet.cell(row=base_row, column=column).value)
                    if re.match(group_name_pattern, cell):
                        groups_schedule[cell] = ScheduleTable()
                        for row in range(base_row + 2, last_row + 1):
                            subject = str(
                                sheet.cell(row=row, column=column).value
                            )
                            if "\n" in subject:
                                subject = subject[: subject.index("\n")]
                            position = groups_schedule[cell].append(subject)

                            teachers = re.findall(
                                teacher_name_pattern,
                                str(
                                    sheet.cell(
                                        row=row, column=column + 2
                                    ).value
                                ),
                            )
                            for teacher in teachers:
                                if teacher not in teachers_schedule.keys():
                                    teachers_schedule[
                                        teacher
                                    ] = ScheduleTable()
                                teachers_schedule[teacher].add(
                                    *position, subject
                                )

        for group in groups_schedule:
            pickle.dump(
                groups_schedule[group],
                open(
                    os.path.join(self.pickle_tables_dir, group + ".pkl"), "wb"
                ),
            )

        for teacher in teachers_schedule:
            pickle.dump(
                teachers_schedule[teacher],
                open(
                    os.path.join(self.pickle_tables_dir, teacher + ".pkl"),
                    "wb",
                ),
            )

    def get_week_number(self, day=None):
        if day is None:
            day = datetime.date.today()

        if day > datetime.date(day.year, 9, 1):
            return (
                day.isocalendar()[1]
                - datetime.date(day.year, 9, 1).isocalendar()[1]
            )
        else:
            return (
                day.isocalendar()[1]
                - datetime.date(day.year, 2, 1).isocalendar()[1]
            )

    def get_week_parity(self, day=None):
        if day is None:
            day = datetime.date.today()

        week_num = self.get_week_number(day)
        if week_num % 2 == 0:
            return Parity.even
        else:
            return Parity.odd

    def get_groups_list(self):
        res = []
        group_name_pattern = r"^[А-Я]{4}-\d{2}-\d{2}$"
        for filename in os.listdir(self.pickle_tables_dir):
            if re.match(group_name_pattern, filename[:-4]):
                res.append(filename[:-4])
        return res

    def get_teachers_list(self):
        res = []
        teacher_name_pattern = r"[А-ЯЁ][а-яё]+\s[А-Я]\.[А-Я]\."
        for filename in os.listdir(self.pickle_tables_dir):
            if re.match(teacher_name_pattern, filename[:-4]):
                res.append(filename[:-4])
        return res

    def get_schedules(self, name):
        res = dict()
        for filename in os.listdir(self.pickle_tables_dir):
            if filename == name or filename[: len(name)] == name:
                path = os.path.join(self.pickle_tables_dir, filename)
                schedule = pickle.load(open(path, "rb"))
                res[filename[:-4]] = schedule
        return res

    def get_formatted_general_schedules(self, name):
        res = dict()
        schedules = self.get_schedules(name)
        for found_name in schedules:
            buff = f"Расписание занятий {found_name}:\n"

            for weekday in Weekdays:
                if weekday == Weekdays.sunday:
                    break
                buff += "\n" + str(weekday) + ":\n"
                for pair in range(1, schedules[found_name].pairs_per_day + 1):
                    buff += f"\n{pair})"
                    buff += f"\nI  - {schedules[found_name].get(weekday, pair, Parity.odd)}".replace(
                        "None", ""
                    )
                    buff += f"\nII - {schedules[found_name].get(weekday, pair, Parity.even)}\n".replace(
                        "None", ""
                    )

            res[found_name] = buff
        return res

    def get_formatted_general_day_schedules(self, name, weekday):
        res = dict()
        schedules = self.get_schedules(name)
        for found_name in schedules:
            buff = f"Расписание занятий {found_name} на {weekday}:"
            for pair in range(1, schedules[found_name].pairs_per_day + 1):
                buff += f"\n{pair})"
                buff += f"\nI  - {schedules[found_name].get(weekday, pair, Parity.odd)}".replace(
                    "None", ""
                )
                buff += f"\nII - {schedules[found_name].get(weekday, pair, Parity.even)}\n".replace(
                    "None", ""
                )

            res[found_name] = buff
        return res

    def get_formatted_week_schedules(self, name, day):
        res = dict()

        week_start = day - datetime.timedelta(days=day.weekday())
        for days_from_start in range(0, 6):
            weekday = week_start + datetime.timedelta(days=days_from_start)
            weekday_schedules = self.get_formatted_day_schedules(name, weekday)
            if weekday == week_start:
                res = weekday_schedules
            else:
                for found_name in weekday_schedules:
                    res[found_name] += "\n\n" + weekday_schedules[found_name]

        return res

    def get_formatted_day_schedules(self, name, day):
        res = dict()
        schedules = self.get_schedules(name)
        for found_name in schedules:
            week_parity = self.get_week_parity(day)
            weekday = Weekdays(day.isocalendar()[2])

            buff = (
                f"Расписание занятий {found_name} на {day.strftime('%d.%m')}:"
            )
            for pair in range(1, schedules[found_name].pairs_per_day + 1):
                buff += f"\n{pair}) {schedules[found_name].get(weekday, pair, week_parity)}".replace(
                    "None", ""
                )

            res[found_name] = buff
        return res

    def run(self):
        while not self.stopped:
            print("Schedule: updating started")
            t1 = time.time()
            self.update_schedule()
            t2 = time.time()
            print(f"Schedule: updating ended ({round(t2 - t1, 1)}s)")
            time.sleep(self.update_time)
        print("Schedule: stopped")

    def stop(self):
        self.stopped = True
